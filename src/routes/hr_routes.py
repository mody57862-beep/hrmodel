from flask import Blueprint, request, jsonify, send_file
from src.models.hr_models import db, Employee, LeaveManagement, LeaveRequest, Attendance, Department, Document
from datetime import datetime, date
import sqlite3
import os
import io
from werkzeug.utils import secure_filename

# استيراد openpyxl بدلاً من pandas
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

hr_bp = Blueprint("hr", __name__)

# مسار للحصول على جميع الموظفين
@hr_bp.route("/employees", methods=["GET"])
def get_employees():
    try:
        employees = Employee.query.all()
        return jsonify([emp.to_dict() for emp in employees])
    except Exception as e:
        return jsonify({"error": str(e)}), 500

# مسار للحصول على موظف محدد
@hr_bp.route("/employees/<int:employee_id>", methods=["GET"])
def get_employee(employee_id):
    try:
        employee = Employee.query.get_or_404(employee_id)
        return jsonify(employee.to_dict())
    except Exception as e:
        return jsonify({"error": str(e)}), 500

# مسار لإضافة موظف جديد
@hr_bp.route("/employees", methods=["POST"])
def add_employee():
    try:
        data = request.get_json()

        # تحويل التواريخ من نص إلى كائن date
        hire_date = datetime.strptime(data.get("hire_date"), "%Y-%m-%d").date() if data.get("hire_date") else None
        birth_date = datetime.strptime(data.get("birth_date"), "%Y-%m-%d").date() if data.get("birth_date") else None
        contract_end_date = datetime.strptime(data.get("contract_end_date"), "%Y-%m-%d").date() if data.get("contract_end_date") else None

        employee = Employee(
            employee_id=data.get("employee_id"),
            full_name=data.get("full_name"),
            house_number=data.get("house_number"),
            national_id=data.get("national_id"),
            job_title=data.get("job_title"),
            qualification=data.get("qualification"),
            hire_date=hire_date,
            points_count=data.get("points_count", 0),
            years_of_experience=data.get("years_of_experience", 0),
            salary_from_system=data.get("salary_from_system"),
            actual_salary=data.get("actual_salary"),
            department_code=data.get("department_code"),
            department=data.get("department"),  # للتوافق المؤقت
            email=data.get("email"),
            phone=data.get("phone"),
            birth_date=birth_date,
            nationality=data.get("nationality"),
            id_number=data.get("id_number"),
            address=data.get("address"),
            marital_status=data.get("marital_status"),
            children_count=data.get("children_count", 0),
            education_level=data.get("education_level"),
            specialization=data.get("specialization"),
            contract_end_date=contract_end_date,
            basic_salary=data.get("basic_salary"),
            allowances=data.get("allowances"),
            total_salary=data.get("total_salary"),
            bank_account=data.get("bank_account"),
            notes=data.get("notes"),
        )

        db.session.add(employee)
        db.session.commit()

        return jsonify(employee.to_dict()), 201
    except Exception as e:
        db.session.rollback()
        return jsonify({"error": str(e)}), 500

# مسار لتحديث بيانات موظف
@hr_bp.route("/employees/<int:employee_id>", methods=["PUT"])
def update_employee(employee_id):
    try:
        employee = Employee.query.get_or_404(employee_id)
        data = request.get_json()

        # تحديث البيانات
        for key, value in data.items():
            if hasattr(employee, key):
                if key in ["hire_date", "birth_date", "contract_end_date"] and value:
                    value = datetime.strptime(value, "%Y-%m-%d").date()
                setattr(employee, key, value)

        employee.updated_at = datetime.utcnow()
        db.session.commit()

        return jsonify(employee.to_dict())
    except Exception as e:
        db.session.rollback()
        return jsonify({"error": str(e)}), 500

# مسار لحذف موظف
@hr_bp.route("/employees/<int:employee_id>", methods=["DELETE"])
def delete_employee(employee_id):
    try:
        employee = Employee.query.get_or_404(employee_id)

        # حذف السجلات المرتبطة أولاً
        LeaveManagement.query.filter_by(employee_id=employee_id).delete()
        LeaveRequest.query.filter_by(employee_id=employee_id).delete()
        Attendance.query.filter_by(employee_id=employee_id).delete()

        # حذف الموظف
        db.session.delete(employee)
        db.session.commit()

        return jsonify({"message": "تم حذف الموظف بنجاح"}), 200
    except Exception as e:
        db.session.rollback()
        return jsonify({"error": str(e)}), 500

# مسار لتصدير بيانات الموظفين إلى Excel باستخدام openpyxl
@hr_bp.route("/employees/export", methods=["GET"])
def export_employees_to_excel():
    try:
        employees = Employee.query.all()
        
        # إنشاء مصنف جديد
        wb = Workbook()
        ws = wb.active
        ws.title = "بيانات الموظفين"
        
        # تعريف رؤوس الأعمدة بالترتيب المطلوب
        headers = [
            'الرقم الوظيفي', 'الاسم الكامل', 'الدار', 'الرقم القومي', 'الوظيفة', 'المؤهل',
            'تاريخ التعيين', 'عدد الابناط', 'سنوات الخبرة', 'الراتب من المنظومة', 'الراتب',
            'كود القسم', 'القسم', 'البريد الإلكتروني', 'رقم الهاتف', 'تاريخ الميلاد',
            'الجنسية', 'العنوان', 'الحالة الاجتماعية', 'عدد الأطفال', 'المستوى التعليمي',
            'التخصص', 'تاريخ انتهاء العقد', 'الراتب الأساسي', 'البدلات', 'إجمالي الراتب',
            'رقم الحساب البنكي', 'ملاحظات'
        ]
        ws.append(headers)
        
        # تنسيق رؤوس الأعمدة
        header_font = Font(bold=True)
        for cell in ws[1]:
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')

        # إضافة البيانات
        for emp in employees:
            emp_dict = emp.to_dict()
            row_data = [
                emp_dict.get('employee_id'),
                emp_dict.get('full_name'),
                emp_dict.get('house_number'),
                emp_dict.get('national_id'),
                emp_dict.get('job_title'),
                emp_dict.get('qualification'),
                emp_dict.get('hire_date'),
                emp_dict.get('points_count'),
                emp_dict.get('years_of_experience'),
                emp_dict.get('salary_from_system'),
                emp_dict.get('actual_salary'),
                emp_dict.get('department_code'),
                emp_dict.get('department'),
                emp_dict.get('email'),
                emp_dict.get('phone'),
                emp_dict.get('birth_date'),
                emp_dict.get('nationality'),
                emp_dict.get('address'),
                emp_dict.get('marital_status'),
                emp_dict.get('children_count'),
                emp_dict.get('education_level'),
                emp_dict.get('specialization'),
                emp_dict.get('contract_end_date'),
                emp_dict.get('basic_salary'),
                emp_dict.get('allowances'),
                emp_dict.get('total_salary'),
                emp_dict.get('bank_account'),
                emp_dict.get('notes')
            ]
            ws.append(row_data)

        # ضبط عرض الأعمدة تلقائياً
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter # Get the column name
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2 # إضافة هامش بسيط
            ws.column_dimensions[column].width = adjusted_width

        # حفظ المصنف في الذاكرة
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # إنشاء اسم الملف مع التاريخ
        filename = f"employees_data_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        return jsonify({"error": f"خطأ في تصدير البيانات: {str(e)}"}), 500

# مسار لاستيراد بيانات الموظفين من Excel باستخدام openpyxl
@hr_bp.route("/employees/import", methods=["POST"])
def import_employees_from_excel():
    try:
        if 'file' not in request.files:
            return jsonify({"error": "لم يتم رفع أي ملف"}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({"error": "لم يتم اختيار ملف"}), 400
        
        if not file.filename.lower().endswith(('.xlsx', '.xls')):
            return jsonify({"error": "يجب أن يكون الملف من نوع Excel (.xlsx أو .xls)"}), 400
        
        # تحميل المصنف
        try:
            wb = load_workbook(file)
            ws = wb.active
        except Exception as e:
            return jsonify({"error": f"خطأ في قراءة ملف Excel: {str(e)}"}), 400
        
        # تعريف رؤوس الأعمدة المتوقعة باللغة العربية ومقابلها في قاعدة البيانات
        column_mapping = {
            'الرقم الوظيفي': 'employee_id',
            'الاسم الكامل': 'full_name',
            'الدار': 'house_number',
            'الرقم القومي': 'national_id',
            'الوظيفة': 'job_title',
            'المؤهل': 'qualification',
            'تاريخ التعيين': 'hire_date',
            'عدد الابناط': 'points_count',
            'سنوات الخبرة': 'years_of_experience',
            'الراتب من المنظومة': 'salary_from_system',
            'الراتب': 'actual_salary',
            'كود القسم': 'department_code',
            'القسم': 'department',
            'البريد الإلكتروني': 'email',
            'رقم الهاتف': 'phone',
            'تاريخ الميلاد': 'birth_date',
            'الجنسية': 'nationality',
            'العنوان': 'address',
            'الحالة الاجتماعية': 'marital_status',
            'عدد الأطفال': 'children_count',
            'المستوى التعليمي': 'education_level',
            'التخصص': 'specialization',
            'تاريخ انتهاء العقد': 'contract_end_date',
            'الراتب الأساسي': 'basic_salary',
            'البدلات': 'allowances',
            'إجمالي الراتب': 'total_salary',
            'رقم الحساب البنكي': 'bank_account',
            'ملاحظات': 'notes'
        }
        
        # قراءة رؤوس الأعمدة من الصف الأول في ملف Excel
        excel_headers = [cell.value for cell in ws[1]]
        
        # إنشاء قائمة بالأسماء الإنجليزية للأعمدة بناءً على الترتيب في Excel
        mapped_headers = []
        for header in excel_headers:
            mapped_headers.append(column_mapping.get(header, None)) # استخدم None إذا لم يتم العثور على تطابق

        imported_count = 0
        updated_count = 0
        errors = []
        
        # قراءة البيانات بدءاً من الصف الثاني
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            employee_data = {}
            for col_idx, cell_value in enumerate(row):
                if col_idx < len(mapped_headers) and mapped_headers[col_idx]:
                    employee_data[mapped_headers[col_idx]] = cell_value
            
            try:
                employee_id = employee_data.get('employee_id')
                full_name = employee_data.get('full_name')

                if not employee_id:
                    errors.append(f"الصف {row_idx}: الرقم الوظيفي مطلوب.")
                    continue
                if not full_name:
                    errors.append(f"الصف {row_idx}: الاسم الكامل مطلوب.")
                    continue
                
                # تحويل الأنواع
                employee_data['employee_id'] = int(employee_id)
                employee_data['points_count'] = int(employee_data.get('points_count', 0))
                employee_data['years_of_experience'] = int(employee_data.get('years_of_experience', 0))
                employee_data['children_count'] = int(employee_data.get('children_count', 0))
                
                # تحويل التواريخ
                date_fields = ['hire_date', 'birth_date', 'contract_end_date']
                for field in date_fields:
                    if field in employee_data and employee_data[field]:
                        if isinstance(employee_data[field], datetime):
                            employee_data[field] = employee_data[field].date()
                        elif isinstance(employee_data[field], str):
                            try:
                                employee_data[field] = datetime.strptime(employee_data[field], "%Y-%m-%d").date()
                            except ValueError:
                                employee_data[field] = None # أو التعامل مع الخطأ بشكل آخر
                        else:
                            employee_data[field] = None
                    else:
                        employee_data[field] = None
                
                # تحويل الأرقام العشرية
                float_fields = ['salary_from_system', 'actual_salary', 'basic_salary', 'allowances', 'total_salary']
                for field in float_fields:
                    if field in employee_data and employee_data[field] is not None:
                        try:
                            employee_data[field] = float(employee_data[field])
                        except ValueError:
                            employee_data[field] = None

                existing_employee = Employee.query.filter_by(employee_id=employee_data['employee_id']).first()
                
                if existing_employee:
                    for key, value in employee_data.items():
                        if key != 'employee_id':
                            setattr(existing_employee, key, value)
                    existing_employee.updated_at = datetime.utcnow()
                    updated_count += 1
                else:
                    new_employee = Employee(**employee_data)
                    db.session.add(new_employee)
                    imported_count += 1
                
            except Exception as e:
                errors.append(f"الصف {row_idx}: {str(e)}")
                continue
        
        db.session.commit()
        
        result = {
            "message": "تم استيراد البيانات بنجاح",
            "imported_count": imported_count,
            "updated_count": updated_count,
            "total_processed": imported_count + updated_count,
            "errors_count": len(errors)
        }
        
        if errors:
            result["errors"] = errors[:10]  # عرض أول 10 أخطاء فقط
            if len(errors) > 10:
                result["note"] = f"تم عرض أول 10 أخطاء من أصل {len(errors)} خطأ"
        
        return jsonify(result), 200
        
    except Exception as e:
        db.session.rollback()
        return jsonify({"error": f"خطأ عام في استيراد البيانات: {str(e)}"}), 500

# مسار للحصول على إدارة الإجازات
@hr_bp.route("/leave-management", methods=["GET"])
def get_leave_management():
    try:
        leave_records = LeaveManagement.query.all()
        return jsonify([record.to_dict() for record in leave_records])
    except Exception as e:
        return jsonify({"error": str(e)}), 500

# مسار للحصول على إدارة الإجازات لموظف محدد
@hr_bp.route("/leave-management/<int:employee_id>", methods=["GET"])
def get_employee_leave_management(employee_id):
    try:
        leave_record = LeaveManagement.query.filter_by(employee_id=employee_id).first()
        if leave_record:
            return jsonify(leave_record.to_dict())
        else:
            return jsonify({"error": "Leave management record not found"}), 404
    except Exception as e:
        return jsonify({"error": str(e)}), 500

# مسار للحصول على طلبات الإجازات
@hr_bp.route("/leave-requests", methods=["GET"])
def get_leave_requests():
    try:
        leave_requests = LeaveRequest.query.all()
        return jsonify([req.to_dict() for req in leave_requests])
    except Exception as e:
        return jsonify({"error": str(e)}), 500

# مسار لإضافة طلب إجازة جديد
@hr_bp.route("/leave-requests", methods=["POST"])
def add_leave_request():
    try:
        data = request.get_json()

        start_date = datetime.strptime(data.get("start_date"), "%Y-%m-%d").date()
        end_date = datetime.strptime(data.get("end_date"), "%Y-%m-%d").date()

        leave_request = LeaveRequest(
            employee_id=data.get("employee_id"),
            leave_type=data.get("leave_type"),
            start_date=start_date,
            end_date=end_date,
            days_requested=data.get("days_requested"),
            reason=data.get("reason"),
            status=data.get("status", "pending"),
        )

        db.session.add(leave_request)
        db.session.commit()

        return jsonify(leave_request.to_dict()), 201
    except Exception as e:
        db.session.rollback()
        return jsonify({"error": str(e)}), 500

# مسار للحصول على سجلات الحضور
@hr_bp.route("/attendance", methods=["GET"])
def get_attendance():
    try:
        attendance_records = Attendance.query.all()
        return jsonify([record.to_dict() for record in attendance_records])
    except Exception as e:
        return jsonify({"error": str(e)}), 500

# مسار للحصول على سجلات الحضور لموظف محدد
@hr_bp.route("/attendance/<int:employee_id>", methods=["GET"])
def get_employee_attendance(employee_id):
    try:
        attendance_records = Attendance.query.filter_by(employee_id=employee_id).all()
        return jsonify([record.to_dict() for record in attendance_records])
    except Exception as e:
        return jsonify({"error": str(e)}), 500

# مسار للحصول على الأقسام
@hr_bp.route("/departments", methods=["GET"])
def get_departments():
    try:
        departments = Department.query.all()
        return jsonify([dept.to_dict() for dept in departments])
    except Exception as e:
        return jsonify({"error": str(e)}), 500

# مسار للحصول على إحصائيات عامة
@hr_bp.route("/statistics", methods=["GET"])
def get_statistics():
    try:
        # استخدام SQLite مباشرة للاستعلامات المعقدة
        db_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), "database", "hr_system.db")
        conn = sqlite3.connect(db_path)
        cursor = conn.cursor()

        # إحصائيات الموظفين
        cursor.execute("SELECT COUNT(*) FROM employees")
        total_employees = cursor.fetchone()[0]

        cursor.execute("SELECT department, COUNT(*) FROM employees GROUP BY department")
        employees_by_department = dict(cursor.fetchall())

        # إحصائيات الحضور
        cursor.execute("SELECT AVG(working_hours) FROM attendance WHERE working_hours IS NOT NULL")
        avg_working_hours = cursor.fetchone()[0] or 0

        cursor.execute("SELECT AVG(late_minutes) FROM attendance WHERE late_minutes IS NOT NULL")
        avg_late_minutes = cursor.fetchone()[0] or 0

        # إحصائيات الإجازات
        cursor.execute("SELECT SUM(annual_leave_used), SUM(casual_leave_used), SUM(sick_leave_used) FROM leave_management")
        leave_stats = cursor.fetchone()

        conn.close()

        statistics = {
            "total_employees": total_employees,
            "employees_by_department": employees_by_department,
            "avg_working_hours": round(avg_working_hours, 2),
            "avg_late_minutes": round(avg_late_minutes, 2),
            "total_annual_leave_used": leave_stats[0] or 0,
            "total_casual_leave_used": leave_stats[1] or 0,
            "total_sick_leave_used": leave_stats[2] or 0,
        }

        return jsonify(statistics)
    except Exception as e:
        return jsonify({"error": str(e)}), 500

# مسار للبحث في الموظفين
@hr_bp.route("/employees/search", methods=["GET"])
def search_employees():
    try:
        query = request.args.get("q", "")
        department = request.args.get("department", "")

        employees_query = Employee.query

        if query:
            employees_query = employees_query.filter(
                Employee.full_name.contains(query) | Employee.employee_id.like(f"%{query}%")
            )

        if department:
            employees_query = employees_query.filter(Employee.department == department)

        employees = employees_query.all()
        return jsonify([emp.to_dict() for emp in employees])
    except Exception as e:
        return jsonify({"error": str(e)}), 500

# مسار للحصول على جميع الأقسام (للقوائم المنسدلة)
@hr_bp.route("/departments_list", methods=["GET"])
def get_departments_list():
    try:
        departments = Department.query.with_entities(Department.name).all()
        return jsonify([dept[0] for dept in departments])
    except Exception as e:
        return jsonify({"error": str(e)}), 500

# مسار للحصول على الموظفين حسب القسم (للقوائم المنسدلة)
@hr_bp.route("/employees_by_department/<string:department_name>", methods=["GET"])
def get_employees_by_department(department_name):
    try:
        employees = Employee.query.filter_by(department=department_name).all()
        return jsonify([{"employee_id": emp.employee_id, "full_name": emp.full_name} for emp in employees])
    except Exception as e:
        return jsonify({"error": str(e)}), 500
